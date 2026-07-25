# ==========================================
# 1. ΟΛΑ ΤΑ IMPORTS ΕΔΩ (Μία φορά)
# ==========================================
import cv2
import numpy as np
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import customtkinter as ctk
from tkinterdnd2 import TkinterDnD, DND_FILES
from tkinter import filedialog, messagebox
import threading 
import datetime  
import platform  
import subprocess
from PIL import Image, ImageTk
import rawpy  
import time
import winsound
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

# Ρύθμιση Light Theme
ctk.set_appearance_mode("Light")  

# ==========================================
# 2. ΚΛΑΣΗ 1ΗΣ ΕΦΑΡΜΟΓΗΣ (Ανάλυση)
# ==========================================
class BananaAnalysisApp(TkinterDnD.Tk):
    def __init__(self):
        super().__init__()
        
        self.title("Banana App — Advanced RAW Colorimetry Platform")
        # --- ΕΞΥΠΝΟ ΚΕΝΤΡΑΡΙΣΜΑ ΓΙΑ ΤΗΝ ΑΝΑΛΥΣΗ ---
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        
        # Επιλέγει το μικρότερο: είτε το ιδανικό πλάτος/ύψος, είτε το 90% της οθόνης σου
        w = min(1450, int(screen_width * 0.9))
        h = min(1020, int(screen_height * 0.9))
        
        x = int((screen_width / 2) - (w / 2))
        y = max(0, int((screen_height / 2) - (h / 2))) # Το max(0, ...) αποτρέπει να βγει το παράθυρο πάνω από την οθόνη
        
        self.geometry(f"{w}x{h}+{x}+{y}")  
        self.resizable(True, True)
        self.minsize(1200, 850)
        self.configure(background="#FEF9C3")  # Ζεστό Απαλό Μπανανί Φόντο
        
        # Cursor Μαϊμούς
        monkey_cursor_path = "C:/Users/Ειρήνη/Downloads/monkey.cur"
        if os.path.exists(monkey_cursor_path):
            try:
                self.configure(cursor=f"@{monkey_cursor_path}")
            except Exception:
                pass
        
        try:
            self.iconbitmap("C:/Users/Ειρήνη/Downloads/app_icon.ico")
        except Exception:
            pass
        
        # --- ΠΡΟΣΘΗΚΗ: Αλλαγή συμπεριφοράς του κουμπιού X ---
        self.protocol("WM_DELETE_WINDOW", self.back_to_menu)

        self.current_img_lab = None      
        self.current_banana_mask = None  
        self.current_summary_data = None 
        self.current_plot_path = None    
        self.display_scale = 1.0         
        self.display_width = 0
        self.display_height = 0
        
        # === 1. ΑΡΙΣΤΕΡΟ PANEL (SIDEBAR) ===
        self.sidebar_frame = ctk.CTkFrame(self, width=460, corner_radius=0, fg_color="#FFFBEB")
        self.sidebar_frame.pack(side="left", fill="y")
        self.sidebar_frame.pack_propagate(False)
        
        # Header
        self.title_label = ctk.CTkLabel(
            self.sidebar_frame, 
            text="🍌 Banana App Pro", 
            font=ctk.CTkFont(family="Segoe UI", size=30, weight="bold"),
            text_color="#D97706"
        )
        self.title_label.pack(pady=(20, 2), padx=25, anchor="w")
        
        self.subtitle_label = ctk.CTkLabel(
            self.sidebar_frame, 
            text="Advanced CIELAB & Ripeness Analytics", 
            font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
            text_color="#78350F"
        )
        self.subtitle_label.pack(pady=(0, 10), padx=25, anchor="w")
        
        # === CONTROLS CONTAINER ===
        self.controls_frame = ctk.CTkFrame(self.sidebar_frame, fg_color="#FEF08A", corner_radius=10, border_color="#F59E0B", border_width=2)
        self.controls_frame.pack(fill="x", padx=15, pady=(0, 10))
        
        # Κουμπιά Επιλογής
        self.btn_browse_single = ctk.CTkButton(
            self.controls_frame, 
            text="📷 Επιλογή Μίας Εικόνας (NEF)", 
            command=self.browse_single_file, 
            font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"),
            height=40,
            fg_color="#D97706",
            hover_color="#B45309",
            text_color="#FFFFFF"
        )
        self.btn_browse_single.pack(fill="x", padx=12, pady=(10, 6))
        
        self.btn_browse_batch = ctk.CTkButton(
            self.controls_frame, 
            text="📁 Μαζική Επεξεργασία Φακέλου", 
            command=self.browse_batch_folder, 
            font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"),
            height=40,
            fg_color="#B45309",
            hover_color="#78350F",
            text_color="#FFFFFF"
        )
        self.btn_browse_batch.pack(fill="x", padx=12, pady=(0, 10))
        
        # === ΑΝΕΞΑΡΤΗΤΕΣ ΕΠΙΛΟΓΕΣ ΕΞΑΓΩΓΗΣ ===
        self.output_label = ctk.CTkLabel(self.controls_frame, text="📦 Επιλογές Εξαγωγής Αρχείων:", font=ctk.CTkFont(size=12, weight="bold"), text_color="#78350F")
        self.output_label.pack(padx=12, anchor="w", pady=(0, 2))

        self.chk_summary_var = ctk.BooleanVar(value=True)
        self.chk_summary = ctk.CTkCheckBox(
            self.controls_frame, text="📊 Συγκεντρωτική Καρτέλα Στατιστικών", 
            variable=self.chk_summary_var, font=ctk.CTkFont(size=11, weight="bold"), 
            text_color="#78350F", fg_color="#D97706"
        )
        self.chk_summary.pack(anchor="w", padx=12, pady=(2, 4))

        self.chk_pixels_var = ctk.BooleanVar(value=True)
        self.chk_pixels = ctk.CTkCheckBox(
            self.controls_frame, text="🔍 Αναλυτικά Pixels ανά Slice", 
            variable=self.chk_pixels_var, font=ctk.CTkFont(size=11, weight="bold"), 
            text_color="#78350F", fg_color="#D97706"
        )
        self.chk_pixels.pack(anchor="w", padx=12, pady=(0, 4))

        self.chk_plots_var = ctk.BooleanVar(value=True)
        self.chk_plots = ctk.CTkCheckBox(
            self.controls_frame, text="📈 Διαγράμματα (Plots)", 
            variable=self.chk_plots_var, font=ctk.CTkFont(size=11, weight="bold"), 
            text_color="#78350F", fg_color="#D97706"
        )
        self.chk_plots.pack(anchor="w", padx=12, pady=(0, 10))

        # === ΕΠΙΛΟΓΗ ΔΕΙΓΜΑΤΟΛΗΨΙΑΣ PIXELS ===
        self.sampling_label = ctk.CTkLabel(self.controls_frame, text="⚙️ Καταγραφή Pixels στο Excel:", font=ctk.CTkFont(size=12, weight="bold"), text_color="#78350F")
        self.sampling_label.pack(padx=12, anchor="w", pady=(2, 2))

        self.sampling_mode = ctk.CTkSegmentedButton(
            self.controls_frame,
            values=["Όλα τα Pixels (100%)", "1 στα 5 Pixels (20%)"],
            font=ctk.CTkFont(size=11, weight="bold"),
            selected_color="#D97706", unselected_color="#FFFBEB", text_color="#451A03"
        )
        self.sampling_mode.set("1 στα 5 Pixels (20%)")
        self.sampling_mode.pack(fill="x", padx=12, pady=(0, 10))

        # Toggles
        self.show_labels_var = ctk.BooleanVar(value=True)
        self.chk_labels = ctk.CTkCheckBox(
            self.controls_frame, text="Εμφάνιση Labels (Slice_1, κτλ) στην εικόνα", 
            variable=self.show_labels_var, font=ctk.CTkFont(size=11, weight="bold"), 
            text_color="#78350F", fg_color="#D97706"
        )
        self.chk_labels.pack(anchor="w", padx=12, pady=(5, 10))

        # Progress Area
        self.progress_label = ctk.CTkLabel(self.sidebar_frame, text="Αναμονή αρχείου...", font=ctk.CTkFont(size=12, weight="bold"), text_color="#B45309")
        self.progress_label.pack(pady=(2, 2), padx=25, anchor="w")

        self.progress_bar = ctk.CTkProgressBar(self.sidebar_frame, height=10, progress_color="#D97706", fg_color="#FDE68A")
        self.progress_bar.pack(fill="x", padx=20, pady=(0, 10))
        self.progress_bar.set(0)

        # Logs Box
        self.textbox_label = ctk.CTkLabel(self.sidebar_frame, text="📋 Διαγνωστικά Μηνύματα / Logs:", font=ctk.CTkFont(size=13, weight="bold"), text_color="#78350F")
        self.textbox_label.pack(padx=25, anchor="w", pady=(0, 2))

        self.textbox = ctk.CTkTextbox(self.sidebar_frame, font=ctk.CTkFont(family="Consolas", size=11, weight="bold"), fg_color="#FEF9C3", text_color="#451A03", border_color="#F59E0B", border_width=2, corner_radius=8)
        self.textbox.pack(fill="both", expand=True, padx=20, pady=(0, 15))
        self.textbox.insert("0.0", "Έτοιμο για ανάλυση. Εισάγετε ένα αρχείο .NEF για έναρξη.")
        self.textbox.configure(state="disabled")

        # === 2. ΔΕΞΙ PANEL (MAIN VIEWPORT ΜΕ 2 ΚΑΡΤΕΛΕΣ: ΧΑΡΤΗΣ & ΠΙΝΑΚΑΣ) ===
        self.main_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="#FEF9C3")
        self.main_frame.pack(side="right", fill="both", expand=True)

        self.map_title_label = ctk.CTkLabel(self.main_frame, text="Προεπισκόπηση & Αναλυτικά Δεδομένα", font=ctk.CTkFont(family="Segoe UI", size=20, weight="bold"), text_color="#78350F")
        self.map_title_label.pack(pady=(15, 5))

        # === TAB VIEW (2 Καρτέλες) ===
        self.tab_view = ctk.CTkTabview(
            self.main_frame, 
            fg_color="#FFFBEB", 
            segmented_button_selected_color="#D97706",
            segmented_button_selected_hover_color="#B45309",
            segmented_button_unselected_color="#FEF08A",
            segmented_button_unselected_hover_color="#FDE68A",
            text_color="#451A03",
            command=self.on_tab_changed
        )
        self.tab_view.pack(pady=5, padx=20, fill="both", expand=True)

        self.tab_map = self.tab_view.add("🗺️ Οπτικός Χάρτης & Ιστόγραμμα")
        self.tab_stats = self.tab_view.add("📊 Πίνακας Στατιστικών")

        # --- 1. TAB: Οπτικός Χάρτης + Σταγονόμετρο + Ιστόγραμμα ---
        self.map_scroll_frame = ctk.CTkScrollableFrame(self.tab_map, fg_color="#FFFBEB", corner_radius=8)
        self.map_scroll_frame.pack(fill="both", expand=True, padx=5, pady=5)

        # Εικόνα / Visual Map
        self.image_container = ctk.CTkFrame(self.map_scroll_frame, fg_color="#FFFBEB", corner_radius=12, border_color="#F59E0B", border_width=2)
        self.image_container.pack(fill="x", padx=5, pady=5)

        self.image_label = ctk.CTkLabel(self.image_container, text="📥\n\nΣύρετε (Drag & Drop) εικόνες .NEF εδώ\nή χρησιμοποιήστε τα κουμπιά επιλογής", fg_color="transparent", text_color="#92400E", font=ctk.CTkFont(size=16, weight="bold"))  
        self.image_label.pack(fill="both", expand=True, padx=10, pady=20)
        self.image_label.bind("<Motion>", self.on_mouse_move)

        # Σταγονόμετρο
        self.eyedropper_frame = ctk.CTkFrame(self.map_scroll_frame, height=75, corner_radius=10, fg_color="#FFFBEB", border_color="#F59E0B", border_width=2)
        self.eyedropper_frame.pack(pady=8, fill="x", padx=5)
        self.eyedropper_frame.pack_propagate(False)

        self.eyedropper_title = ctk.CTkLabel(self.eyedropper_frame, text="🧪 ΣΤΑΓΟΝΟΜΕΤΡΟ (REAL-TIME CIELAB EYEDROPPER)", font=ctk.CTkFont(size=11, weight="bold"), text_color="#D97706")
        self.eyedropper_title.pack(anchor="w", padx=15, pady=(6, 0))

        self.eyedropper_value_label = ctk.CTkLabel(self.eyedropper_frame, text="Μετακινήστε το ποντίκι πάνω από τις φέτες για ζωντανή μέτρηση.", font=ctk.CTkFont(family="Consolas", size=13, weight="bold"), text_color="#451A03")
        self.eyedropper_value_label.pack(anchor="w", padx=15, pady=(0, 6))

        # Ιστόγραμμα & Plots
        self.plot_container = ctk.CTkFrame(self.map_scroll_frame, fg_color="#FFFBEB", corner_radius=12, border_color="#F59E0B", border_width=2)
        self.plot_container.pack(fill="x", padx=5, pady=(5, 10))

        self.plot_title_label = ctk.CTkLabel(self.plot_container, text="📈 Ιστόγραμμα & Καταγραφή Χρωμάτων (Plots)", font=ctk.CTkFont(size=12, weight="bold"), text_color="#D97706")
        self.plot_title_label.pack(anchor="w", padx=15, pady=(8, 0))

        self.plot_image_label = ctk.CTkLabel(self.plot_container, text="Τα διαγράμματα θα εμφανιστούν εδώ μετά την ανάλυση.", fg_color="transparent", text_color="#92400E", font=ctk.CTkFont(size=14, weight="bold"))
        self.plot_image_label.pack(fill="x", expand=True, padx=10, pady=10)

        # --- 2. TAB: Πίνακας Στατιστικών ---
        self.stats_textbox = ctk.CTkTextbox(
            self.tab_stats, 
            font=ctk.CTkFont(family="Consolas", size=13, weight="bold"), 
            fg_color="#FFFBEB", 
            text_color="#451A03",
            border_color="#F59E0B", 
            border_width=2,
            corner_radius=8
        )
        self.stats_textbox.pack(fill="both", expand=True, padx=5, pady=5)
        self.stats_textbox.insert("0.0", "Τα στατιστικά στοιχεία θα εμφανιστούν εδώ μετά την ανάλυση.")
        self.stats_textbox.configure(state="disabled")

        try:
            self.drop_target_register(DND_FILES)
            self.dnd_bind('<<Drop>>', self.on_file_drop)
        except Exception:
            pass

    # --- ΠΡΟΣΘΗΚΗ: Συνάρτηση Επιστροφής στο Μενού ---
    def back_to_menu(self):
        self.destroy()
        launcher = BananaLauncher()
        launcher.mainloop()

    def on_tab_changed(self):
        current_tab = self.tab_view.get()
        if current_tab == "🗺️ Οπτικός Χάρτης & Ιστόγραμμα":
            # Χρησιμοποιούμε το before=self.plot_container για να μπει στη μέση
            self.eyedropper_frame.pack(before=self.plot_container, pady=8, fill="x", padx=5)
        else:
            self.eyedropper_frame.pack_forget()

    def on_file_drop(self, event):
        paths = self.parse_drop_paths(event.data)
        nef_files = [p for p in paths if p.lower().endswith('.nef')]
        if nef_files:
            self.start_batch_or_single(nef_files)
        else:
            messagebox.showwarning("Μη έγκυρο αρχείο", "Παρακαλώ σύρετε αρχεία .NEF")

    def parse_drop_paths(self, data):
        import re
        return [g[0] or g[1] for g in re.findall(r'\{([^}]+)\}|(\S+)', data)]

    def browse_single_file(self):
        f = filedialog.askopenfilename(filetypes=[("Nikon RAW Images", "*.nef;*.NEF")])
        if f: self.start_batch_or_single([f])

    def browse_batch_folder(self):
        folder = filedialog.askdirectory(title="Επιλέξτε φάκελο με αρχεία NEF")
        if folder:
            nef_files = [os.path.join(folder, f) for f in os.listdir(folder) if f.lower().endswith('.nef')]
            if nef_files:
                self.start_batch_or_single(nef_files)
            else:
                messagebox.showinfo("Δεν βρέθηκαν αρχεία", "Δεν βρέθηκαν αρχεία .NEF στον φάκελο.")

    def start_batch_or_single(self, file_list):
        if not (self.chk_summary_var.get() or self.chk_pixels_var.get() or self.chk_plots_var.get()):
            messagebox.showwarning("Επιλογή Εξαγωγής", "Παρακαλώ επιλέξτε τουλάχιστον μία επιλογή εξαγωγής.")
            return

        default_dir = os.path.dirname(file_list[0])
        selected_parent_dir = filedialog.askdirectory(
            title="Επιλέξτε φάκελο αποθήκευσης των αποτελεσμάτων", 
            initialdir=default_dir
        )
        if not selected_parent_dir: 
            return

        self.btn_browse_single.configure(state="disabled")
        self.btn_browse_batch.configure(state="disabled")
        self.image_label.configure(image=None, text="⏳ Γίνεται επεξεργασία...")
        self.image_label.image = None

        step = 5 if "1 στα 5" in self.sampling_mode.get() else 1
        show_labels = self.show_labels_var.get()

        opts = {
            'summary': self.chk_summary_var.get(),
            'pixels': self.chk_pixels_var.get(),
            'plots': self.chk_plots_var.get()
        }

        self.current_img_lab = None
        self.current_banana_mask = None

        threading.Thread(target=self.process_batch, args=(file_list, selected_parent_dir, step, show_labels, opts), daemon=True).start()

    def log_message(self, message):
        self.textbox.configure(state="normal")
        self.textbox.insert("end", message + "\n")
        self.textbox.configure(state="disabled")
        self.textbox.see("end")

    def update_progress(self, value, text_override=None):
        self.progress_bar.set(value)
        if text_override: self.progress_label.configure(text=text_override)
        else: self.progress_label.configure(text=f"Πρόοδος: {int(value * 100)}%")
        self.update_idletasks()

    def open_directory(self, path):
        if platform.system() == "Windows": 
            os.startfile(path)

    def display_map_image(self, img_path):
        try:
            pil_img = Image.open(img_path)
            max_height = 420
            self.display_scale = max_height / float(pil_img.size[1])
            self.display_width = int(float(pil_img.size[0]) * float(self.display_scale))
            self.display_height = max_height
            pil_img = pil_img.resize((self.display_width, self.display_height), Image.Resampling.LANCZOS)
            ctk_img = ctk.CTkImage(light_image=pil_img, dark_image=pil_img, size=(self.display_width, self.display_height))
            
            self.image_label.configure(image=ctk_img, text="")
            self.image_label.image = ctk_img  
        except Exception as e:
            self.log_message(f"⚠️ Αδυναμία προβολής εικόνας: {str(e)}")

    def display_plot_image(self, plot_path):
        try:
            pil_img = Image.open(plot_path)
            max_height = 380
            w = int(float(pil_img.size[0]) * (max_height / float(pil_img.size[1])))
            pil_img = pil_img.resize((w, max_height), Image.Resampling.LANCZOS)
            ctk_img = ctk.CTkImage(light_image=pil_img, dark_image=pil_img, size=(w, max_height))
            
            self.plot_image_label.configure(image=ctk_img, text="")
            self.plot_image_label.image = ctk_img
        except Exception as e:
            self.log_message(f"⚠️ Αδυναμία προβολής διαγράμματος: {str(e)}")

    def update_stats_view(self, summary_rows):
        text_out = "ΣΥΓΚΕΝΤΡΩΤΙΚΑ ΣΤΑΤΙΣΤΙΚΑ (CIELAB & RIPENESS)\n"
        text_out += "="*95 + "\n"
        text_out += f"{'Slice':<10} | {'Pixels':<8} | {'L* Mean':<9} | {'a* Mean':<9} | {'b* Mean':<9} | {'Chroma (C*)':<12} | {'Hue (h°)':<9}\n"
        text_out += "-"*95 + "\n"
        
        for r in summary_rows:
            text_out += f"{r[0]:<10} | {r[1]:<8} | {r[2]:<9} | {r[6]:<9} | {r[8]:<9} | {r[10]:<12} | {r[11]:<9}\n"
        
        text_out += "="*95 + "\n"
        
        self.stats_textbox.configure(state="normal")
        self.stats_textbox.delete("0.0", "end")
        self.stats_textbox.insert("0.0", text_out)
        self.stats_textbox.configure(state="disabled")

    def on_mouse_move(self, event):
        if self.current_img_lab is None or self.current_banana_mask is None: return  
        x_display, y_display = event.x, event.y
        if 0 <= x_display < self.display_width and 0 <= y_display < self.display_height:
            x_real = min(int(x_display / self.display_scale), self.current_img_lab.shape[1] - 1)
            y_real = min(int(y_display / self.display_scale), self.current_img_lab.shape[0] - 1)
            
            if self.current_banana_mask[y_real, x_real] > 0:
                L, a, b = self.current_img_lab[y_real, x_real]
                C = np.sqrt(a**2 + b**2)
                h = np.degrees(np.arctan2(b, a)) % 360
                self.eyedropper_value_label.configure(
                    text=f"📍 X: {x_real:,}  Y: {y_real:,}   │   L*: {L:.2f}   │   a*: {a:.2f}   │   b*: {b:.2f}   │   Chroma: {C:.2f}   │   Hue: {h:.1f}°",
                    text_color="#B45309"
                )
            else:
                self.eyedropper_value_label.configure(
                    text=f"📍 X: {x_real:,}  Y: {y_real:,}   │   [Περιοχή Background]",
                    text_color="#78350F"
                )

    def process_batch(self, file_list, destination_base_dir, sample_step, show_labels, opts):
        total_files = len(file_list)
        self.textbox.configure(state="normal")
        self.textbox.delete("0.0", "end")
        self.textbox.configure(state="disabled")
        
        self.log_message(f"🚀 Έναρξη επεξεργασίας {total_files} αρχείων...\n")

        last_target_dir = None
        for f_idx, input_path in enumerate(file_list, 1):
            self.log_message(f"=== ({f_idx}/{total_files}) Επεξεργασία: {os.path.basename(input_path)} ===")
            target_dir = self.process_single_image(input_path, destination_base_dir, sample_step, show_labels, opts, f_idx, total_files)
            if target_dir:
                last_target_dir = target_dir

        self.update_progress(1.0, "Ολοκληρώθηκε!")
        self.log_message("\n🎉 ΟΛΑ ΤΑ ΑΡΧΕΙΑ ΕΠΕΞΕΡΓΑΣΤΗΚΑΝ ΜΕ ΕΠΙΤΥΧΙΑ!")
        self.eyedropper_value_label.configure(text="🧪 Σταγονόμετρο Ενεργό! Κουνήστε το ποντίκι πάνω από την εικόνα.", text_color="#B45309")
        
        try: winsound.MessageBeep(winsound.MB_ICONASTERISK)
        except Exception: pass
        
        messagebox.showinfo("Ολοκλήρωση", f"Η επεξεργασία ολοκληρώθηκε επιτυχώς!")
        if last_target_dir:
            self.open_directory(last_target_dir)
        
        self.btn_browse_single.configure(state="normal")
        self.btn_browse_batch.configure(state="normal")

    def process_single_image(self, input_path, destination_base_dir, sample_step, show_labels, opts, f_idx=1, total_files=1):
        file_name = os.path.splitext(os.path.basename(input_path))[0]
        prefix = "1στα5_Pixels" if sample_step == 5 else "Όλα_Pixels"
        
        folder_name = f"Αποτελέσματα_{prefix}_{file_name}"
        target_dir = os.path.join(destination_base_dir, folder_name)

        if os.path.exists(target_dir):
            answer = messagebox.askyesno(
                "Ο φάκελος υπάρχει ήδη", 
                f"Ο φάκελος '{folder_name}' υπάρχει ήδη στον επιλεγμένο προορισμό.\n\n"
                "• Πατήστε 'Ναι' για ΑΝΤΙΚΑΤΑΣΤΑΣΗ των αρχείων.\n"
                "• Πατήστε 'Όχι' για ΔΗΜΙΟΥΡΓΙΑ ΝΕΟΥ φακέλου."
            )
            if not answer: 
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                target_dir = os.path.join(destination_base_dir, f"Αποτελέσματα_{prefix}_{file_name}_{timestamp}")
                os.makedirs(target_dir, exist_ok=True)
        else:
            os.makedirs(target_dir, exist_ok=True)

        output_xlsx_summary = os.path.join(target_dir, f'{file_name}_summary_analysis.xlsx')
        output_xlsx_pixels = os.path.join(target_dir, f'{file_name}_pixels_analysis.xlsx')
        output_img = os.path.join(target_dir, f'{file_name}_visual_map.png')
        output_plot = os.path.join(target_dir, f'{file_name}_color_plots.png')

        self.log_message(f"📂 Αρχείο RAW: {input_path}")
        self.log_message(f"📁 Φάκελος εξόδου: {target_dir}")
        self.log_message("⏳ Ανάπτυξη και φόρτωση RAW...")
        
        start_time = time.time()
        self.update_progress(0.10, f"({f_idx}/{total_files}) Ανάπτυξη RAW (10%)")

        try:
            with rawpy.imread(input_path) as raw:
                img_rgb = raw.postprocess(use_camera_wb=True) 
            
            img_bgr = cv2.cvtColor(img_rgb, cv2.COLOR_RGB2BGR)
            gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
            blurred = cv2.GaussianBlur(gray, (5, 5), 0)

            thresh_val, banana_mask = cv2.threshold(blurred, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            self.log_message(f"💡 Κατώφλι Otsu: {int(thresh_val)}")
            self.log_message("🔍 Εντοπισμός δειγμάτων...")
            self.update_progress(0.25, f"({f_idx}/{total_files}) Εντοπισμός δειγμάτων (25%)")

            mask_floodfill = banana_mask.copy()
            h_m, w_m = banana_mask.shape[:2]
            floodfill_mask = np.zeros((h_m + 2, w_m + 2), np.uint8)
            cv2.floodFill(mask_floodfill, floodfill_mask, (0,0), 255)
            banana_mask = banana_mask | cv2.bitwise_not(mask_floodfill)

            num_labels, labels, stats, centroids = cv2.connectedComponentsWithStats(banana_mask)
            valid_objects = [(i, centroids[i], stats[i, cv2.CC_STAT_AREA]) for i in range(1, num_labels) if stats[i, cv2.CC_STAT_AREA] >= 5000]

            if len(valid_objects) < 6:
                self.log_message(f"❌ Σφάλμα: Βρέθηκαν μόνο {len(valid_objects)} αντικείμενα (απαιτούνται 6). Παράλειψη.")
                return

            valid_objects.sort(key=lambda x: x[2], reverse=True)
            valid_objects = valid_objects[:6]
            valid_objects.sort(key=lambda x: x[1][1]) 
            ordered_objects = sorted(valid_objects[:3], key=lambda x: x[1][0]) + sorted(valid_objects[3:6], key=lambda x: x[1][0])

            clean_mask = np.zeros_like(banana_mask)
            for obj_id, _, _ in ordered_objects: clean_mask[labels == obj_id] = 255
            self.current_banana_mask = clean_mask

            bg_color = np.array([70, 55, 45], dtype=np.uint8)
            img_bgr_cleaned = np.where(clean_mask[:, :, None] == 255, img_bgr, bg_color)
            
            img_float = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB).astype(np.float32) / 255.0
            img_lab = cv2.cvtColor(img_float, cv2.COLOR_RGB2Lab)
            self.current_img_lab = img_lab

            output_map = img_bgr_cleaned.copy()
            slices_data = {}
            summary_rows = []

            self.log_message("📊 Υπολογισμός τιμών CIELAB...")
            self.update_progress(0.40, f"({f_idx}/{total_files}) Υπολογισμός CIELAB (40%)")

            for idx, (obj_id, (cX, cY), _) in enumerate(ordered_objects):
                slice_name = f"Slice_{idx + 1}"
                obj_mask = (labels == obj_id)
                
                L_all = img_lab[:, :, 0][obj_mask]
                a_all = img_lab[:, :, 1][obj_mask]
                b_all = img_lab[:, :, 2][obj_mask]

                L_mean, L_std, L_min, L_max = np.mean(L_all), np.std(L_all), np.min(L_all), np.max(L_all)
                a_mean, a_std = np.mean(a_all), np.std(a_all)
                b_mean, b_std = np.mean(b_all), np.std(b_all)
                Chroma_mean = np.sqrt(a_mean**2 + b_mean**2)
                Hue_mean = np.degrees(np.arctan2(b_mean, a_mean)) % 360

                summary_rows.append([
                    slice_name, len(L_all), 
                    round(float(L_mean), 2), round(float(L_std), 2), round(float(L_min), 2), round(float(L_max), 2),
                    round(float(a_mean), 2), round(float(a_std), 2),
                    round(float(b_mean), 2), round(float(b_std), 2),
                    round(float(Chroma_mean), 2), round(float(Hue_mean), 2)
                ])

                if opts['pixels'] or opts['plots']:
                    L_sub, a_sub, b_sub = L_all[::sample_step], a_all[::sample_step], b_all[::sample_step]
                    slices_data[slice_name] = np.column_stack((np.arange(1, len(L_sub) + 1), L_sub, a_sub, b_sub))

                if show_labels:
                    ys, xs = np.where(obj_mask)
                    min_y = np.min(ys)
                    label_x = int(cX) - 130
                    label_y = max(min_y - 40, 50)
                    cv2.putText(output_map, slice_name, (label_x, label_y), cv2.FONT_HERSHEY_SIMPLEX, 2.5, (0, 0, 0), 12)
                    cv2.putText(output_map, slice_name, (label_x, label_y), cv2.FONT_HERSHEY_SIMPLEX, 2.5, (255, 255, 255), 5)

            _, im_buf = cv2.imencode('.png', output_map)
            im_buf.tofile(output_img)

            # Ενημέρωση UI: Visual Map & Stats Table
            self.after(0, lambda p=output_img: self.display_map_image(p))
            self.after(0, lambda r=summary_rows: self.update_stats_view(r))

            if opts['plots'] and slices_data:
                self.log_message("📈 Δημιουργία διαγραμμάτων...")
                self.update_progress(0.60, f"({f_idx}/{total_files}) Δημιουργία Plots (60%)")
                
                fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(10, 4.5))
                fig.patch.set_facecolor('#FFFBEB')

                for s_name, data in slices_data.items():
                    ax1.scatter(data[:, 2], data[:, 3], s=1, alpha=0.3, label=s_name)
                    ax2.hist(data[:, 1], bins=30, alpha=0.5, label=s_name)

                ax1.set_title("Color Space (a* vs b*)", fontsize=10, fontweight='bold')
                ax1.set_xlabel("a* (Green -> Red)", fontsize=9)
                ax1.set_ylabel("b* (Blue -> Yellow)", fontsize=9)
                ax1.grid(True, linestyle='--', alpha=0.5)

                ax2.set_title("Lightness (L*) Histogram", fontsize=10, fontweight='bold')
                ax2.set_xlabel("L* Value", fontsize=9)
                ax2.set_ylabel("Pixel Count", fontsize=9)
                ax2.grid(True, linestyle='--', alpha=0.5)
                ax2.legend(loc='upper right', fontsize=7)

                plt.tight_layout()
                plt.savefig(output_plot, dpi=160, bbox_inches='tight')
                plt.close()

                self.after(0, lambda plt_p=output_plot: self.display_plot_image(plt_p))

            # === ΕΓΓΡΑΦΗ EXCEL ΣΕ ΞΕΧΩΡΙΣΤΑ ΑΡΧΕΙΑ ===
            self.log_message("📝 Εγγραφή αρχείων Excel...")
            
            if opts['summary']:
                wb_sum = openpyxl.Workbook()
                ws_sum = wb_sum.active
                ws_sum.title = "Summary_Stats"
                ws_sum.views.sheetView[0].showGridLines = True
                headers_summary = ['Slice', 'Total_Pixels', 'L*_Mean', 'L*_StdDev', 'L*_Min', 'L*_Max', 'a*_Mean', 'a*_StdDev', 'b*_Mean', 'b*_StdDev', 'Chroma_C*', 'Hue_Angle_h°']
                ws_sum.append(headers_summary)
                for r in summary_rows: ws_sum.append(r)
                wb_sum.save(output_xlsx_summary)
                self.log_message("  📥 Αρχείο 'summary_analysis.xlsx' δημιουργήθηκε.")

            if opts['pixels']:
                wb_px = openpyxl.Workbook()
                wb_px.remove(wb_px.active)
                num_slices = len(slices_data)
                start_time_excel = time.time()

                for s_idx, (slice_key, pixel_matrix) in enumerate(slices_data.items(), 1):
                    ws_px = wb_px.create_sheet(title=slice_key)
                    ws_px.views.sheetView[0].showGridLines = True
                    ws_px.append(['Pixel_Index', 'L*', 'a*', 'b*'])
                    for row_data in pixel_matrix:
                        ws_px.append([int(row_data[0]), float(row_data[1]), float(row_data[2]), float(row_data[3])])

                    elapsed = time.time() - start_time_excel
                    time_per_slice = elapsed / s_idx
                    slices_left = num_slices - s_idx
                    remaining_seconds = int(time_per_slice * slices_left)
                    time_text = f"Απομένουν: {remaining_seconds}δ" if remaining_seconds > 0 else "Σχεδόν έτοιμο..."

                    progress_frac = 0.70 + (s_idx / num_slices) * 0.25
                    self.update_progress(progress_frac, f"Εγγραφή: {slice_key} ({int(progress_frac * 100)}%) | {time_text}")
                    self.log_message(f"  📥 Καρτέλα '{slice_key}' ολοκληρώθηκε.")

                wb_px.save(output_xlsx_pixels)
                self.log_message("  📥 Αρχείο 'pixels_analysis.xlsx' δημιουργήθηκε.")

            self.update_progress(0.95, f"({f_idx}/{total_files}) Οριστικοποίηση (95%)")
            
            total_duration = time.time() - start_time
            self.log_message(f"✔️ Ολοκληρώθηκε σε {total_duration:.1f} δευτερόλεπτα: {file_name}\n")
            return target_dir

        except Exception as e:
            self.log_message(f"❌ Σφάλμα στο {file_name}: {str(e)}\n")
            return None


# ==========================================
# 3. ΚΛΑΣΗ 2ΗΣ ΕΦΑΡΜΟΓΗΣ (Σύγκριση)
# ==========================================
class BananaCompareApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        self.title("Banana App — Εργαλείο Σύγκρισης Δειγμάτων (Delta E)")
        
        # --- ΕΞΥΠΝΟ ΚΕΝΤΡΑΡΙΣΜΑ ΓΙΑ ΤΗ ΣΥΓΚΡΙΣΗ ---
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        
        w = min(1500, int(screen_width * 0.9))
        h = min(950, int(screen_height * 0.9))
        
        x = int((screen_width / 2) - (w / 2))
        y = max(0, int((screen_height / 2) - (h / 2)))
        
        self.geometry(f"{w}x{h}+{x}+{y}")
        self.minsize(1200, 800)
        self.configure(background="#FEF9C3") 
        
        # --- ΠΡΟΣΘΗΚΗ: Αλλαγή συμπεριφοράς του κουμπιού X ---
        self.protocol("WM_DELETE_WINDOW", self.back_to_menu)

        self.file1_path = None
        self.file2_path = None
        
        # ==========================================
        # 1. ΑΡΙΣΤΕΡΗ ΣΤΗΛΗ (CONTROLS & INPUTS)
        # ==========================================
        self.left_panel = ctk.CTkFrame(self, width=460, fg_color="transparent")
        self.left_panel.pack(side="left", fill="y", padx=(20, 10), pady=20)
        self.left_panel.pack_propagate(False)

        # --- HEADER ---
        self.header_frame = ctk.CTkFrame(self.left_panel, fg_color="#FFFBEB", corner_radius=10, border_color="#F59E0B", border_width=2)
        self.header_frame.pack(fill="x", pady=(0, 15))
        
        self.title_label = ctk.CTkLabel(
            self.header_frame, 
            text="⚖️ Σύγκριση Ωρίμανσης", 
            font=ctk.CTkFont(family="Segoe UI", size=22, weight="bold"),
            text_color="#D97706"
        )
        self.title_label.pack(pady=(15, 5))
        
        self.subtitle_label = ctk.CTkLabel(
            self.header_frame, 
            text="Ρυθμίσεις & Επιλογή Αρχείων", 
            font=ctk.CTkFont(family="Segoe UI", size=13),
            text_color="#78350F"
        )
        self.subtitle_label.pack(pady=(0, 15))

        # -- ΑΡΧΙΚΟ ΔΕΙΓΜΑ (T1) --
        self.frame_t1 = ctk.CTkFrame(self.left_panel, fg_color="#FFFBEB", corner_radius=10, border_color="#D97706", border_width=1)
        self.frame_t1.pack(fill="x", pady=(0, 15))

        ctk.CTkLabel(self.frame_t1, text="🕒 1. ΑΡΧΙΚΟ ΔΕΙΓΜΑ (Πριν)", font=ctk.CTkFont(size=15, weight="bold"), text_color="#B45309").pack(pady=(10, 5))
        
        self.btn_file1 = ctk.CTkButton(self.frame_t1, text="Επιλογή Αρχείου .NEF", command=self.select_file1, fg_color="#D97706", hover_color="#B45309")
        self.btn_file1.pack(pady=5)
        self.lbl_file1 = ctk.CTkLabel(self.frame_t1, text="Δεν έχει επιλεγεί αρχείο", text_color="#78350F", font=ctk.CTkFont(size=12, slant="italic"))
        self.lbl_file1.pack(pady=(0, 5))

        self.entry_name1 = ctk.CTkEntry(self.frame_t1, placeholder_text="π.χ. Μπανάνα 1", width=300)
        self.entry_name1.pack(pady=5)
        ctk.CTkLabel(self.frame_t1, text="Όνομα/Περιγραφή Δείγματος", font=ctk.CTkFont(size=11), text_color="#78350F").pack(pady=(0, 5))

        self.entry_time1 = ctk.CTkEntry(self.frame_t1, placeholder_text="π.χ. 0 Ώρες (Αφετηρία)", width=300)
        self.entry_time1.pack(pady=5)
        ctk.CTkLabel(self.frame_t1, text="Χρόνος Λήψης", font=ctk.CTkFont(size=11), text_color="#78350F").pack(pady=(0, 10))

        # -- ΤΕΛΙΚΟ ΔΕΙΓΜΑ (T2) --
        self.frame_t2 = ctk.CTkFrame(self.left_panel, fg_color="#FFFBEB", corner_radius=10, border_color="#059669", border_width=1)
        self.frame_t2.pack(fill="x", pady=(0, 15))

        ctk.CTkLabel(self.frame_t2, text="🏁 2. ΤΕΛΙΚΟ ΔΕΙΓΜΑ (Μετά)", font=ctk.CTkFont(size=15, weight="bold"), text_color="#047857").pack(pady=(10, 5))
        
        self.btn_file2 = ctk.CTkButton(self.frame_t2, text="Επιλογή Αρχείου .NEF", command=self.select_file2, fg_color="#059669", hover_color="#047857")
        self.btn_file2.pack(pady=5)
        self.lbl_file2 = ctk.CTkLabel(self.frame_t2, text="Δεν έχει επιλεγεί αρχείο", text_color="#78350F", font=ctk.CTkFont(size=12, slant="italic"))
        self.lbl_file2.pack(pady=(0, 5))

        self.entry_name2 = ctk.CTkEntry(self.frame_t2, placeholder_text="π.χ. Μπανάνα 1 (Μετά)", width=300)
        self.entry_name2.pack(pady=5)
        ctk.CTkLabel(self.frame_t2, text="Όνομα/Περιγραφή Δείγματος", font=ctk.CTkFont(size=11), text_color="#78350F").pack(pady=(0, 5))

        self.entry_time2 = ctk.CTkEntry(self.frame_t2, placeholder_text="π.χ. 48 Ώρες", width=300)
        self.entry_time2.pack(pady=5)
        ctk.CTkLabel(self.frame_t2, text="Χρόνος Λήψης", font=ctk.CTkFont(size=11), text_color="#78350F").pack(pady=(0, 10))

        # --- ΕΠΙΛΟΓΕΣ ΔΙΑΓΡΑΜΜΑΤΩΝ ---
        self.options_frame = ctk.CTkFrame(self.left_panel, fg_color="#FEF08A", corner_radius=8)
        self.options_frame.pack(fill="x", pady=(0, 15))
        
        ctk.CTkLabel(self.options_frame, text="📊 Διαγράμματα προς Εξαγωγή:", font=ctk.CTkFont(size=13, weight="bold"), text_color="#78350F").pack(anchor="w", padx=15, pady=(10, 5))
        
        self.chk_plot1_var = ctk.BooleanVar(value=True)
        self.chk_plot1 = ctk.CTkCheckBox(self.options_frame, text="Μετατόπιση a* vs b*", variable=self.chk_plot1_var, fg_color="#D97706", text_color="#451A03")
        self.chk_plot1.pack(anchor="w", padx=25, pady=2)

        self.chk_plot2_var = ctk.BooleanVar(value=True)
        self.chk_plot2 = ctk.CTkCheckBox(self.options_frame, text="Ανάλυση ΔL*, Δa*, Δb*", variable=self.chk_plot2_var, fg_color="#D97706", text_color="#451A03")
        self.chk_plot2.pack(anchor="w", padx=25, pady=2)

        self.chk_plot3_var = ctk.BooleanVar(value=True)
        self.chk_plot3 = ctk.CTkCheckBox(self.options_frame, text="Συνολική Απόκλιση ΔE*", variable=self.chk_plot3_var, fg_color="#D97706", text_color="#451A03")
        self.chk_plot3.pack(anchor="w", padx=25, pady=(2, 10))

        # --- ΕΚΚΙΝΗΣΗ ---
        self.btn_compare = ctk.CTkButton(
            self.left_panel, 
            text="🚀 Επιλογή Φακέλου & Έναρξη", 
            command=self.start_comparison, 
            font=ctk.CTkFont(family="Segoe UI", size=16, weight="bold"),
            height=50,
            fg_color="#2563EB",
            hover_color="#1D4ED8",
            text_color="#FFFFFF"
        )
        self.btn_compare.pack(fill="x", pady=10)

        # --- PROGRESS ---
        self.progress_label = ctk.CTkLabel(self.left_panel, text="Αναμονή...", font=ctk.CTkFont(size=12, weight="bold"), text_color="#B45309")
        self.progress_label.pack(pady=(5, 2))

        self.progress_bar = ctk.CTkProgressBar(self.left_panel, height=12, progress_color="#D97706", fg_color="#FDE68A")
        self.progress_bar.pack(fill="x", pady=(0, 10))
        self.progress_bar.set(0)

        # ==========================================
        # 2. ΔΕΞΙΑ ΣΤΗΛΗ (ΑΠΟΤΕΛΕΣΜΑΤΑ & ΓΡΑΦΗΜΑΤΑ)
        # ==========================================
        self.right_panel = ctk.CTkFrame(self, fg_color="transparent")
        self.right_panel.pack(side="right", fill="both", expand=True, padx=(10, 20), pady=20)

        self.tab_view = ctk.CTkTabview(
            self.right_panel, 
            fg_color="#FFFBEB", 
            border_color="#F59E0B", 
            border_width=2,
            segmented_button_selected_color="#D97706",
            segmented_button_selected_hover_color="#B45309"
        )
        self.tab_view.pack(fill="both", expand=True)
        
        self.tab_report = self.tab_view.add("📝 Κείμενο Αναφοράς")
        self.tab_plots = self.tab_view.add("📈 Γραφήματα")
        
        self.result_textbox = ctk.CTkTextbox(
            self.tab_report, 
            font=ctk.CTkFont(family="Consolas", size=16, weight="bold"),
            fg_color="#FFFBEB", 
            text_color="#451A03",
            corner_radius=8
        )
        self.result_textbox.pack(fill="both", expand=True, padx=10, pady=10)
        self.result_textbox.insert("0.0", "Τα αποτελέσματα της σύγκρισης θα εμφανιστούν εδώ.\n\nΟδηγίες:\n1. Επέλεξε το Αρχικό αρχείο.\n2. Επέλεξε το Τελικό αρχείο.\n3. Πάτησε 'Έναρξη Σύγκρισης' για να διαλέξεις φάκελο αποθήκευσης.")
        self.result_textbox.configure(state="disabled")

        self.plots_scroll_frame = ctk.CTkScrollableFrame(self.tab_plots, fg_color="transparent")
        self.plots_scroll_frame.pack(fill="both", expand=True, padx=5, pady=5)
        self.lbl_no_plots = ctk.CTkLabel(self.plots_scroll_frame, text="Δεν έχουν εξαχθεί διαγράμματα ακόμα.", font=ctk.CTkFont(size=16, slant="italic"), text_color="#78350F")
        self.lbl_no_plots.pack(pady=40)

    # --- ΠΡΟΣΘΗΚΗ: Συνάρτηση Επιστροφής στο Μενού ---
    def back_to_menu(self):
        self.destroy()
        launcher = BananaLauncher()
        launcher.mainloop()

    def select_file1(self):
        f = filedialog.askopenfilename(title="Επιλογή Αρχικού Αρχείου (Πριν)", filetypes=[("Nikon RAW Images", "*.nef;*.NEF")])
        if f:
            self.file1_path = f
            self.lbl_file1.configure(text=os.path.basename(f), text_color="#047857")

    def select_file2(self):
        f = filedialog.askopenfilename(title="Επιλογή Τελικού Αρχείου (Μετά)", filetypes=[("Nikon RAW Images", "*.nef;*.NEF")])
        if f:
            self.file2_path = f
            self.lbl_file2.configure(text=os.path.basename(f), text_color="#047857")

    def start_comparison(self):
        if not self.file1_path or not self.file2_path:
            messagebox.showwarning("Ελλιπή Στοιχεία", "Παρακαλώ επιλέξτε και τα δύο αρχεία (Αρχικό και Τελικό) πριν ξεκινήσετε.")
            return

        name1 = self.entry_name1.get() or "T1"
        time1 = self.entry_time1.get() or "0 Ώρες"
        name2 = self.entry_name2.get() or "T2"
        time2 = self.entry_time2.get() or "Τελικό"
            
        # Καθαρισμός ονομάτων για αποφυγή σφαλμάτων στα ονόματα φακέλων (αφαιρούμε περίεργους χαρακτήρες)
        safe_name1 = "".join(c for c in name1 if c.isalnum() or c in (' ', '-', '_')).strip().replace(" ", "_")
        safe_name2 = "".join(c for c in name2 if c.isalnum() or c in (' ', '-', '_')).strip().replace(" ", "_")
        
        target_folder_name = f"Σύγκριση_{safe_name1}_vs_{safe_name2}"

        # Ζητάμε από τον χρήστη τον φάκελο αποθήκευσης
        selected_parent_dir = filedialog.askdirectory(title="Επιλέξτε φάκελο προορισμού για τα αποτελέσματα")
        if not selected_parent_dir:
            return  # Ο χρήστης πάτησε ακύρωση

        full_target_path = os.path.join(selected_parent_dir, target_folder_name)

        # Έλεγχος αν υπάρχει ήδη ο φάκελος
        if os.path.exists(full_target_path):
            answer = messagebox.askyesnocancel(
                "Ο Φάκελος Υπάρχει", 
                f"Ο φάκελος '{target_folder_name}' υπάρχει ήδη στον επιλεγμένο προορισμό!\n\n"
                "• 'Ναι' για ΑΝΤΙΚΑΤΑΣΤΑΣΗ αρχείων.\n"
                "• 'Όχι' για ΔΗΜΙΟΥΡΓΙΑ ΝΕΟΥ (προσθήκη ημερομηνίας).\n"
                "• 'Άκυρο' για ακύρωση της διαδικασίας."
            )
            if answer is None: # Άκυρο
                return
            elif answer is False: # Όχι (Δημιουργία νέου)
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                full_target_path = f"{full_target_path}_{timestamp}"

        # Δημιουργούμε τον φάκελο
        os.makedirs(full_target_path, exist_ok=True)

        self.btn_compare.configure(state="disabled")
        self.result_textbox.configure(state="normal")
        self.result_textbox.delete("0.0", "end")
        self.result_textbox.insert("end", "⏳ Η ανάλυση ξεκίνησε, παρακαλώ περιμένετε...\n")
        self.result_textbox.configure(state="disabled")
        
        for widget in self.plots_scroll_frame.winfo_children():
            widget.destroy()
        ctk.CTkLabel(self.plots_scroll_frame, text="⏳ Γίνεται δημιουργία διαγραμμάτων...", font=ctk.CTkFont(size=16, slant="italic"), text_color="#78350F").pack(pady=40)
        
        self.tab_view.set("📝 Κείμενο Αναφοράς")
        
        # Περνάμε το full_target_path στο thread
        threading.Thread(target=self.run_comparison, args=(self.file1_path, name1, time1, self.file2_path, name2, time2, full_target_path), daemon=True).start()

    def update_progress(self, value, text):
        self.progress_bar.set(value)
        self.progress_label.configure(text=text)
        self.update_idletasks()

    def run_comparison(self, img1_path, name1, time1, img2_path, name2, time2, output_folder):
        self.update_progress(0.1, "Εξαγωγή δεδομένων από το Αρχικό δείγμα (T1)...")
        stats_1 = self._extract_cielab_from_nef(img1_path)
        
        self.update_progress(0.5, "Εξαγωγή δεδομένων από το Τελικό δείγμα (T2)...")
        stats_2 = self._extract_cielab_from_nef(img2_path)
        
        if not stats_1 or not stats_2:
            self.update_progress(0, "Σφάλμα κατά την ανάλυση.")
            self.result_textbox.configure(state="normal")
            self.result_textbox.insert("end", "\n❌ Προέκυψε σφάλμα (π.χ. δεν βρέθηκαν 6 δείγματα στις εικόνες). Η σύγκριση ακυρώθηκε.")
            self.result_textbox.configure(state="disabled")
            self.btn_compare.configure(state="normal")
            return
            
        self.update_progress(0.7, "Υπολογισμός Delta E & Δημιουργία Αναφοράς...")
        
        slices_list = []
        dL_list, da_list, db_list, dE_list = [], [], [], []
        a1_list, b1_list, a2_list, b2_list = [], [], [], []
        
        text_out = "ΑΝΑΦΟΡΑ ΣΥΓΚΡΙΣΗΣ & ΩΡΙΜΑΝΣΗΣ\n"
        text_out += "="*105 + "\n"
        text_out += f"[ΑΡΧΙΚΟ] Αρχείο: {os.path.basename(img1_path)} | Όνομα: {name1} | Χρόνος: {time1}\n"
        text_out += f"[ΤΕΛΙΚΟ] Αρχείο: {os.path.basename(img2_path)} | Όνομα: {name2} | Χρόνος: {time2}\n"
        text_out += "="*105 + "\n"
        text_out += f"{'Slice':<10} | {'Δ L* (Φωτειν.)':<18} | {'Δ a* (Πράσινο->Κόκκινο)':<25} | {'Δ b* (Μπλε->Κίτρινο)':<25} | {'ΔE* (Συνολική)':<15}\n"
        text_out += "-"*105 + "\n"
        
        for slice_name in stats_1.keys():
            if slice_name in stats_2:
                L1, a1, b1 = stats_1[slice_name]
                L2, a2, b2 = stats_2[slice_name]
                
                dL = L2 - L1
                da = a2 - a1
                db = b2 - b1
                dE = np.sqrt(dL**2 + da**2 + db**2)
                
                slices_list.append(slice_name)
                dL_list.append(dL)
                da_list.append(da)
                db_list.append(db)
                dE_list.append(dE)
                a1_list.append(a1)
                b1_list.append(b1)
                a2_list.append(a2)
                b2_list.append(b2)
                
                text_out += f"{slice_name:<10} | {dL:>10.2f}         | {da:>15.2f}            | {db:>15.2f}            | {dE:>15.2f}\n"
        
        text_out += "="*105 + "\n\n"
        text_out += "ΕΠΕΞΗΓΗΣΗ ΑΠΟΤΕΛΕΣΜΑΤΩΝ:\n"
        text_out += "• ΔE*: Η συνολική χρωματική διαφορά. Τιμές > 2.3 είναι συνήθως ορατές.\n"
        text_out += "• ΔL*: Θετικό = Πιο φωτεινό. Αρνητικό = Πιο σκοτεινό.\n"
        text_out += "• Δa*: Θετικό = Έχασε πράσινο (ωρίμασε). Αρνητικό = Έγινε πιο πράσινο.\n"
        text_out += "• Δb*: Θετικό = Έγινε πιο κίτρινο. Αρνητικό = Έχασε κίτρινο.\n"
        
        # Ενημέρωση UI
        self.result_textbox.configure(state="normal")
        self.result_textbox.delete("0.0", "end")
        self.result_textbox.insert("0.0", text_out)
        self.result_textbox.configure(state="disabled")

        # Αποθήκευση της αναφοράς σε .txt στον φάκελο
        try:
            report_path = os.path.join(output_folder, "Αναφορά_Αποτελεσμάτων.txt")
            with open(report_path, "w", encoding="utf-8") as f:
                f.write(text_out)
        except Exception as e:
            print(f"Σφάλμα κατά την αποθήκευση του txt: {e}")

        # --- ΔΗΜΙΟΥΡΓΙΑ ΔΙΑΓΡΑΜΜΑΤΩΝ ---
        do_plot1 = self.chk_plot1_var.get()
        do_plot2 = self.chk_plot2_var.get()
        do_plot3 = self.chk_plot3_var.get()

        generated_plots = [] 

        if do_plot1 or do_plot2 or do_plot3:
            self.update_progress(0.85, "Δημιουργία Γραφημάτων...")

            if do_plot1:
                p1_path = os.path.join(output_folder, "01_Shift_Vector_Plot.png")
                plt.figure(figsize=(10, 8)) 
                plt.title(f"Μετατόπιση Χρώματος (a* vs b*)\nΑπό {time1} προς {time2}", fontweight='bold', fontsize=14)
                plt.xlabel("a* (Πράσινο -> Κόκκινο)", fontsize=12)
                plt.ylabel("b* (Μπλε -> Κίτρινο)", fontsize=12)
                plt.grid(True, linestyle='--', alpha=0.7)
                
                all_a = a1_list + a2_list
                all_b = b1_list + b2_list
                plt.xlim(min(all_a) - 2, max(all_a) + 2)
                plt.ylim(min(all_b) - 2, max(all_b) + 2)

                for i, s in enumerate(slices_list):
                    plt.plot(a1_list[i], b1_list[i], 'bo', markersize=8, label='T1 (Αρχικό)' if i==0 else "")
                    plt.plot(a2_list[i], b2_list[i], 'ro', markersize=8, label='T2 (Τελικό)' if i==0 else "")
                    plt.annotate("", xy=(a2_list[i], b2_list[i]), xytext=(a1_list[i], b1_list[i]), 
                                 arrowprops=dict(arrowstyle="->", color="black", lw=2))
                    plt.text(a1_list[i], b1_list[i] + 0.3, s, fontsize=11, fontweight='bold', alpha=0.8)

                plt.legend(fontsize=12)
                plt.savefig(p1_path, dpi=160, bbox_inches='tight') 
                plt.close()
                generated_plots.append(p1_path)

            if do_plot2:
                p2_path = os.path.join(output_folder, "02_dL_da_db_BarChart.png")
                fig, ax = plt.subplots(figsize=(11, 7))
                x = np.arange(len(slices_list))
                width = 0.25

                ax.bar(x - width, dL_list, width, label='ΔL* (Φωτεινότητα)', color='#9CA3AF')
                ax.bar(x, da_list, width, label='Δa* (Πράσινο->Κόκκινο)', color='#EF4444')
                ax.bar(x + width, db_list, width, label='Δb* (Μπλε->Κίτρινο)', color='#F59E0B')

                ax.set_title(f"Ανάλυση Μεταβολών ΔL*, Δa*, Δb* ανά Δείγμα ({time1} -> {time2})", fontweight='bold', fontsize=14)
                ax.set_xticks(x)
                ax.set_xticklabels(slices_list, fontsize=12)
                ax.axhline(0, color='black', linewidth=1)
                ax.legend(fontsize=11)
                
                plt.savefig(p2_path, dpi=160, bbox_inches='tight')
                plt.close()
                generated_plots.append(p2_path)

            if do_plot3:
                p3_path = os.path.join(output_folder, "03_dE_BarChart.png")
                plt.figure(figsize=(10, 6))
                plt.bar(slices_list, dE_list, color='#8B5CF6')
                plt.axhline(y=2.3, color='r', linestyle='--', linewidth=2, label='Όριο Ορατής Διαφοράς (ΔE=2.3)')
                
                plt.title(f"Συνολική Χρωματική Διαφορά (ΔE*) ανά Δείγμα\nΑπό {time1} προς {time2}", fontweight='bold', fontsize=14)
                plt.ylabel("Τιμή ΔE*", fontsize=12)
                plt.xticks(fontsize=12)
                plt.legend(fontsize=11)
                
                for i, v in enumerate(dE_list):
                    plt.text(i, v + 0.15, f"{v:.2f}", ha='center', fontweight='bold', fontsize=11)

                plt.savefig(p3_path, dpi=160, bbox_inches='tight')
                plt.close()
                generated_plots.append(p3_path)

        # Ενημέρωση UI 
        self.after(0, lambda: self._update_plots_ui(generated_plots))

        self.update_progress(1.0, "Σύγκριση ολοκληρώθηκε!")
        self.btn_compare.configure(state="normal")
        
        # Άνοιγμα του επιλεγμένου φακέλου
        if output_folder and os.path.exists(output_folder):
            if platform.system() == "Windows":
                os.startfile(output_folder)

    def _update_plots_ui(self, plot_paths):
        for widget in self.plots_scroll_frame.winfo_children():
            widget.destroy()

        if not plot_paths:
            ctk.CTkLabel(self.plots_scroll_frame, text="Δεν επιλέχθηκαν διαγράμματα προς εξαγωγή.", font=ctk.CTkFont(size=16, slant="italic"), text_color="#78350F").pack(pady=40)
            return

        for path in plot_paths:
            if os.path.exists(path):
                try:
                    img = Image.open(path)
                    max_width = 850 
                    w_percent = (max_width / float(img.size[0]))
                    h_size = int((float(img.size[1]) * float(w_percent)))
                    img = img.resize((max_width, h_size), Image.Resampling.LANCZOS)
                    
                    ctk_img = ctk.CTkImage(light_image=img, size=(max_width, h_size))
                    
                    lbl = ctk.CTkLabel(self.plots_scroll_frame, text="", image=ctk_img)
                    lbl.image = ctk_img  
                    lbl.pack(pady=20)
                except Exception as e:
                    print(f"Σφάλμα φόρτωσης εικόνας {path}: {e}")

        self.tab_view.set("📈 Γραφήματα")

    def _extract_cielab_from_nef(self, input_path):
        try:
            with rawpy.imread(input_path) as raw:
                img_rgb = raw.postprocess(use_camera_wb=True) 
            
            img_bgr = cv2.cvtColor(img_rgb, cv2.COLOR_RGB2BGR)
            gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
            blurred = cv2.GaussianBlur(gray, (5, 5), 0)

            thresh_val, banana_mask = cv2.threshold(blurred, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            
            mask_floodfill = banana_mask.copy()
            h_m, w_m = banana_mask.shape[:2]
            floodfill_mask = np.zeros((h_m + 2, w_m + 2), np.uint8)
            cv2.floodFill(mask_floodfill, floodfill_mask, (0,0), 255)
            banana_mask = banana_mask | cv2.bitwise_not(mask_floodfill)

            num_labels, labels, stats, centroids = cv2.connectedComponentsWithStats(banana_mask)
            valid_objects = [(i, centroids[i], stats[i, cv2.CC_STAT_AREA]) for i in range(1, num_labels) if stats[i, cv2.CC_STAT_AREA] >= 5000]

            if len(valid_objects) < 6:
                return None

            valid_objects.sort(key=lambda x: x[2], reverse=True)
            valid_objects = valid_objects[:6]
            valid_objects.sort(key=lambda x: x[1][1]) 
            ordered_objects = sorted(valid_objects[:3], key=lambda x: x[1][0]) + sorted(valid_objects[3:6], key=lambda x: x[1][0])

            img_float = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB).astype(np.float32) / 255.0
            img_lab = cv2.cvtColor(img_float, cv2.COLOR_RGB2Lab)

            results = {}
            for idx, (obj_id, _, _) in enumerate(ordered_objects):
                slice_name = f"Slice_{idx + 1}"
                obj_mask = (labels == obj_id)
                
                L_mean = np.mean(img_lab[:, :, 0][obj_mask])
                a_mean = np.mean(img_lab[:, :, 1][obj_mask])
                b_mean = np.mean(img_lab[:, :, 2][obj_mask])
                
                results[slice_name] = (L_mean, a_mean, b_mean)

            return results
        except Exception as e:
            return None

# ==========================================
# 4. ΚΕΝΤΡΙΚΟ ΜΕΝΟΥ (LAUNCHER)
# ==========================================
class BananaLauncher(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        self.title("BANNANA-APP — Κεντρικό Μενού")
        
        # --- ΚΕΝΤΡΑΡΙΣΜΑ ΤΟΥ ΠΑΡΑΘΥΡΟΥ ΣΤΗΝ ΟΘΟΝΗ ---
        window_width = 550
        window_height = 380
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        
        # Υπολογισμός των συντεταγμένων X και Y
        x = int((screen_width / 2) - (window_width / 2))
        y = int((screen_height / 2) - (window_height / 2))
        
        # Εφαρμογή διαστάσεων και θέσης (μορφή: πλάτοςxύψος+x+y)
        self.geometry(f"{window_width}x{window_height}+{x}+{y}")
        # ---------------------------------------------
        
        self.resizable(False, False)
        self.configure(background="#FEF9C3") 
        
        # ... (ο υπόλοιπος κώδικας του μενού παραμένει ως έχει) ...
        
        self.title_label = ctk.CTkLabel(
            self, 
            text="🍌 BANNANA-APP Suite", 
            font=ctk.CTkFont(family="Segoe UI", size=26, weight="bold"), 
            text_color="#D97706"
        )
        self.title_label.pack(pady=(40, 5))
        
        self.subtitle_label = ctk.CTkLabel(
            self, 
            text="Επίλεξε ποιο εργαλείο θέλεις να εκκινήσεις", 
            font=ctk.CTkFont(family="Segoe UI", size=14), 
            text_color="#78350F"
        )
        self.subtitle_label.pack(pady=(0, 40))

        self.btn_analysis = ctk.CTkButton(
            self, 
            text="📊 1. Ανάλυση Δειγμάτων (CIELAB & Στατιστικά)", 
            command=self.run_analysis, 
            height=55, 
            font=ctk.CTkFont(family="Segoe UI", size=15, weight="bold"), 
            fg_color="#D97706", 
            hover_color="#B45309"
        )
        self.btn_analysis.pack(pady=10, padx=60, fill="x")

        self.btn_compare = ctk.CTkButton(
            self, 
            text="⚖️ 2. Σύγκριση Ωρίμανσης (Delta E)", 
            command=self.run_compare, 
            height=55, 
            font=ctk.CTkFont(family="Segoe UI", size=15, weight="bold"), 
            fg_color="#059669", 
            hover_color="#047857"
        )
        self.btn_compare.pack(pady=10, padx=60, fill="x")

    def run_analysis(self):
        self.destroy()  
        app = BananaAnalysisApp()
        app.mainloop()

    def run_compare(self):
        self.destroy()  
        app = BananaCompareApp()
        app.mainloop()

# ==========================================
# 5. ΕΚΚΙΝΗΣΗ ΤΟΥ ΕΝΙΑΙΟΥ ΠΡΟΓΡΑΜΜΑΤΟΣ
# ==========================================
if __name__ == "__main__":
    launcher = BananaLauncher()
    launcher.mainloop()
